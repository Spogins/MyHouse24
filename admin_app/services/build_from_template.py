import openpyxl
from django.db.models import QuerySet
from django.http import JsonResponse
from django.views import View
from django.views.generic.detail import SingleObjectMixin
from django.db.models import Sum, Case, When, Value, FloatField, OuterRef, Subquery, Q, QuerySet
from admin_app.models import Template, Receipt, BankBook, PaymentDetails, ReceiptService
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from myhouse import settings


class BuildReceiptFileView(SingleObjectMixin, View):
    model = Receipt

    def post(self, request, *args, **kwargs):
        try:
            if request.GET.get('template'):
                template = Template.objects.get(pk=request.GET.get('template')).file
            else:
                template = Template.objects.get(is_default=True).file

            wb = openpyxl.Workbook()
            base_template = openpyxl.load_workbook(filename=template.path)
            base_sheet = base_template.active
            receipt = Receipt.objects.get(id=request.GET.get('receipt'))
            banknook = BankBook.objects.get(flat_id=receipt.flat.id)
            payment = PaymentDetails.objects.first()
            reserved_words = {
                '$payCompany$': payment.name if payment else 'N/A',
                '$receiptAddress$': f'{receipt.flat.owner.fullname()}, дом. {receipt.flat.house.name}, кв {receipt.flat.number}.',
                '$total$': receipt.get_price(),
                '$accountBalance$': banknook.balance(),
                '$receiptPayable$': banknook.balance() - receipt.get_price(),
                '$receiptDate$': f'{receipt.date_from}',
                '$receiptMonth$': f'{receipt.date_to}',
                '$accountNumber$': banknook.id,
                '$receiptNumber$': receipt.id,
            }

            from copy import copy

            # copy merged cells from base template
            for merged_range in base_sheet.merged_cells:
                wb.worksheets[0].merge_cells(str(merged_range))

            # copy width and height of each row
            for index, rd in base_sheet.row_dimensions.items():
                wb.worksheets[0].row_dimensions[index] = copy(rd)

            for index, cd in base_sheet.column_dimensions.items():
                wb.worksheets[0].column_dimensions[index] = copy(cd)

            start_moving = None
            end_moving = None
            for index, row in enumerate(base_sheet):
                for index_second, cell in enumerate(row):
                    val = base_sheet.cell(row=index + 1, column=index_second + 1).value

                    # copy styles from parent worksheet
                    if base_sheet.cell(row=index + 1, column=index_second + 1).has_style:
                        wb.worksheets[0].cell(row=index + 1, column=index_second + 1).font = copy(base_sheet.cell(
                            row=index + 1, column=index_second + 1).font)
                        wb.worksheets[0].cell(row=index + 1, column=index_second + 1).border = copy(base_sheet.cell(
                            row=index + 1, column=index_second + 1).border)
                        wb.worksheets[0].cell(row=index + 1, column=index_second + 1).fill = copy(base_sheet.cell(
                            row=index + 1, column=index_second + 1).fill)

                    if val:

                        if val.startswith('$') and val not in ['$serviceName$', '$servicePrice$', '$serviceUnit$', '$serviceAmount$', '$serviceTotal$']:

                            if val == '$LOOP START$':
                                start_moving = base_sheet.cell(row=index + 2, column=index_second + 1)
                                continue
                            if val == '$LOOP ENDING$':
                                end_moving = base_sheet.cell(row=index + 2, column=index_second + 1)
                                continue
                            wb.worksheets[0].cell(row=index + 1, column=index_second + 1).value = reserved_words.get(
                                val)
                            continue

                        wb.worksheets[0].cell(row=index + 1, column=index_second + 1).value = val


            # moving footer of cycle for correct display receipt services
            wb.worksheets[0].move_range(f'{start_moving.coordinate}:{end_moving.coordinate}',
                                        rows=ReceiptService.objects.filter(receipt_id=receipt.id).count() - 3, cols=0)
            current_total_row = start_moving.row + ReceiptService.objects.filter(receipt_id=receipt.id).count() - 3

            # merging cells in rows with services' totals
            wb.worksheets[0].merge_cells(
                f'{get_column_letter(start_moving.column)}{current_total_row}:{get_column_letter(start_moving.column + 1)}{current_total_row}')
            wb.worksheets[0].merge_cells(
                f'{get_column_letter(start_moving.column + 2)}{current_total_row}:{get_column_letter(start_moving.column + 3)}{current_total_row}')
            wb.worksheets[0].merge_cells(
                f'{get_column_letter(start_moving.column + 4)}{current_total_row}:{get_column_letter(start_moving.column + 5)}{current_total_row}')
            wb.worksheets[0].merge_cells(
                f'{get_column_letter(start_moving.column + 6)}{current_total_row}:{get_column_letter(start_moving.column + 7)}{current_total_row}')
            wb.worksheets[0].merge_cells(
                f'{get_column_letter(start_moving.column + 8)}{current_total_row}:{get_column_letter(start_moving.column + 10)}{current_total_row}')

            def set_service_info(name: str, workbook: openpyxl.Workbook, starting_row: int, starting_column: int):
                """Inner function for passing into template service's information"""

                if name == '$serviceName$':
                    receipt_services = [receipt_service.service.name for receipt_service in
                                        ReceiptService.objects.filter(receipt_id=receipt.id)]
                elif name == '$servicePrice$':
                    receipt_services = [receipt_service.price_unit for receipt_service in
                                        ReceiptService.objects.filter(receipt_id=receipt.id)]
                elif name == '$serviceUnit$':
                    receipt_services = [receipt_service.unit.name for receipt_service in ReceiptService.objects.filter(receipt_id=receipt.id)]
                elif name == '$serviceAmount$':
                    receipt_services = [receipt_service.amount for receipt_service in ReceiptService.objects.filter(receipt_id=receipt.id)]
                else:
                    receipt_services = [receipt_service.price for receipt_service in
                                        ReceiptService.objects.filter(receipt_id=receipt.id)]

                for row_index in range(starting_row, starting_row + ReceiptService.objects.filter(receipt_id=receipt.id).count()):

                    # merging cells for better display
                    if name != '$serviceTotal$':
                        workbook.worksheets[0].merge_cells(
                            f'{get_column_letter(starting_column)}{row_index}:{get_column_letter(starting_column + 1)}{row_index}')
                    elif name == '$serviceTotal$':
                        workbook.worksheets[0].merge_cells(
                            f'{get_column_letter(starting_column)}{row_index}:{get_column_letter(starting_column + 2)}{row_index}')

                    # setting for new cells with services styles of previous cells from template
                    if workbook.worksheets[0].cell(row=row_index - 1,
                                                   column=starting_column).has_style and row_index != starting_row:
                        wb.worksheets[0].cell(row=row_index, column=starting_column).font = \
                            copy(workbook.worksheets[0].cell(row=row_index - 1, column=starting_column).font)

                        wb.worksheets[0].cell(row=row_index, column=starting_column).border = \
                            copy(workbook.worksheets[0].cell(row=row_index - 1, column=starting_column).border)

                        wb.worksheets[0].cell(row=row_index, column=starting_column).fill = \
                            copy(workbook.worksheets[0].cell(row=row_index - 1, column=starting_column).fill)

                    workbook.worksheets[0].cell(row=row_index, column=starting_column).value = receipt_services[
                        row_index - starting_row]

            for index, row in enumerate(wb.worksheets[0]):
                for index_second, cell in enumerate(row):
                    val = wb.worksheets[0].cell(row=index + 1, column=index_second + 1).value
                    if val in ['$serviceName$', '$servicePrice$', '$serviceUnit$', '$serviceAmount$', '$serviceTotal$']:
                        set_service_info(val, wb, index + 1, index_second + 1)

            from os import path, mkdir
            if not path.exists(f'{settings.MEDIA_ROOT}/receipts'):
                mkdir(f'{settings.MEDIA_ROOT}/receipts')

            file_name = f'receipt_{receipt.id}.xlsx'
            file_path = f'{settings.MEDIA_ROOT}/receipts/{file_name}'
            wb.save(file_path)
            return JsonResponse(
                {'answer': 'success', 'file_path': f'{settings.MEDIA_URL}receipts/{file_name}', 'full_path': file_path,
                 'file_name': file_name})
        except (Template.DoesNotExist, Receipt.DoesNotExist, BankBook.DoesNotExist) as e:
            return JsonResponse({'answer': 'failed'})

